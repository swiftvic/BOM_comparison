#from openpyxl import Workbook
#from openpyxl import load_workbook
import openpyxl
import datetime

def main():

    filepath1 = "9010242321RAMR0.xlsx"
    old_wb = openpyxl.load_workbook(filepath1)

    filepath2 = "9010242321RA1MR0.xlsx"
    new_wb = openpyxl.load_workbook(filepath2)

    #print(old_wb.sheetnames)

    # Select which sheet in old wb
    old_sheet = old_wb["Sheet1"]

    # Select which sheet in new wb
    new_sheet = new_wb["Sheet1"]

    #print(sheet.max_row)
    #print(sheet.max_column)

    # Old sheet max rows and columns
    max_row = old_sheet.max_row
    max_col = old_sheet.max_column

    # New sheet max rows and columns
    max_new_row = new_sheet.max_row
    max_new_col = new_sheet.max_column

    # Prints stats of each file
    print("There are " + str(max_row) + " line items in " + str(filepath1))
    print("There are " + str(max_new_row) + " line items in " + str(filepath2))

    # Prints out all values on column 3 or C
    '''
    for row in range(1, max_row):
        print(old_sheet.cell(row, 3).value)
    '''

    # Compare column Cs between old and new sheet, prints True if match, False if not plus location of (row, col) and value changed
    '''
    for both_rows in range(1, max_row):
        if old_sheet.cell(both_rows, 3).value == new_sheet.cell(both_rows, 3).value:
            print("True")
            #print(str(both_rows) + ", " + str(3) + " OLD: " + str(old_sheet.cell(both_rows, 3).value) + " NEW: " + str(new_sheet.cell(both_rows,3).value) )
        else:
            print("False")
            print(str(both_rows) + ", " + str(3) + " OLD: " + str(old_sheet.cell(both_rows, 3).value) + " NEW: " + str(new_sheet.cell(both_rows,3).value) )
    '''
    
    # Takes Mara p/n, search through BOM to see if it exists
    # Check old BOM part numbers against new BOM
    not_found = False                                           # Flag for finding value

    for r in range(5, max_row, 2):                              # Loop through rows 
        old_value = old_sheet.cell(r, 2).value

        for new_r in range(5, max_new_row):
            if old_value == new_sheet.cell(new_r, 2).value:
                not_found = False
                break                                           # Found value in new sheet, break out
            else:
                not_found = True
                   
        if not_found:
            print(str(old_value) + " is removed.")

    # Check for new part numbers from new BOM vs old
    new = False           # Flag for finding value

    for rn in range(5, max_new_row, 2):                         # Loop through each row from row 5 new BOM max rows
        part = new_sheet.cell(rn, 2).value

        for old_r in range(5, max_row):                         # Loop through each row from row 5 to max BOM old rows
            if part == old_sheet.cell(old_r, 2).value:          # If found a part number match
                new = False
                break                                           # Break out of for loop
            else:                                               # Cannot find a match
                new = True                                      # Set flag as a new part found
                   
        if new:
            print(str(part) + " is new.")
    
    '''
    for i in range(1, 5):
        for j in range(1, 5):
            print (old_sheet.cell(row=i, column = j).value)
    '''
    
if __name__ == '__main__':

    main()

